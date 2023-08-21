from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title('Data Entery')
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="lightblue")

file=pathlib.Path('backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Phone Number"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"
    
    file.save('backend_data.xlsx')
    
    
def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combo.get()
    address=addressEntry.get(1.0,END)
    
    
    file=openpyxl.load_workbook('backend_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)
    
    file.save(r'backend_data.xlsx')
    
    messagebox.showinfo('info','detail added!')
    
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    gender_combo.set('')
    addressEntry.delete(1.0,END)
    
    
def clear():
    nameValue.set('')
    AgeValue.set('')
    contactValue.set('')
    addressEntry.delete(1.0,END)
# icon
# icon_image=PhotoImage(file="logo.png")
# root.iconphoto(False,icon_image)

# heading
Label(root,text="Please Fill out this Entry Form:",font='arial 13',bg="royalblue",fg="#fff").place(x=20,y=20)

# Label
Label(root,text='Name',font=23,bg='orange',fg="#fff").place(x=50,y=100)
Label(root,text='Contact no.',font=23,bg='orange',fg="#fff").place(x=50,y=150)
Label(root,text='Age',font=23,bg='orange',fg="#fff").place(x=50,y=200)
Label(root,text='Gender',font=23,bg='orange',fg="#fff").place(x=390,y=200)
Label(root,text='Address',font=23,bg='orange',fg="#fff").place(x=50,y=250)

# entry
nameValue=StringVar()
contactValue=StringVar()
AgeValue=StringVar()

nameEntry=Entry(root,textvariable=nameValue,width=37,bd=2,font=20)
conatctEntry=Entry(root,textvariable=contactValue,width=37,bd=2,font=20)
AgeEntry=Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

# gender
gender_combo=Combobox(root,values=['Male','Female','Transender'],font='aria 14',state='r',width=10)
gender_combo.place(x=480,y=200)
gender_combo.set('Male')

addressEntry=Text(root,width=50,height=4,bd=5)

nameEntry.place(x=200,y=100)
conatctEntry.place(x=200,y=150)
AgeEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250) 

# buttons
Button(root,text='Submit',width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text='Clear',width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text='Exit',width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)

root.mainloop()
